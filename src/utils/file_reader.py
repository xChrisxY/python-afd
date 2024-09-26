import os
import pandas as pd
from docx import Document
from bs4 import BeautifulSoup
from openpyxl import load_workbook

def read_text_from_excel(file_path):
    
    content = []
    
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active  

    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            if cell.value is not None: 
                content.append([f"Fila: {cell.row}", f"Columna: {cell.column}", f"Valor: {cell.value}"])
                
    return content

def read_text_from_csv(file_path):
    
    content = []   
    df = pd.read_csv(file_path)
    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row):
            content.append([f"Fila: {row_idx}", f"Columna: {col_idx}", f"Valor: {value}"])
            
    return content

def read_text_from_docx(file_path):
    
    doc = Document(file_path)
    text = '\n'.join(para.text for para in doc.paragraphs)
    return text

def read_text_from_html(file_path):

    with open(file_path, 'r', encoding='utf-8') as file:
        soup = BeautifulSoup(file, 'html.parser')

    text = soup.get_text()
    return text

def read_text_from_txt(file_path):
    
    content = []
    
    with open(file_path, 'r', encoding='utf-8') as file:
        for line in file:
            content.append(line.strip())

    return content

def read_text(file_path):
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()
    if ext == '.csv':
        return read_text_from_csv(file_path)
    elif ext in ['.xls', '.xlsx']:
        return read_text_from_excel(file_path)
    elif ext == '.docx':
        return read_text_from_docx(file_path)
    elif ext in ['.html', '.htm']:
        return read_text_from_html(file_path)
    elif ext == '.txt':
        return read_text_from_txt(file_path)
    else:
        raise ValueError("Formato de archivo no soportado.")

